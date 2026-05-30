"""Generate Nursery_Template.xlsx and Nursery_Hub.xlsx for the proposed workflow.

These workbooks are .xlsx (no embedded VBA) and ship with styled "button" cells
on the Home tab. After import (see SETUP.md), the included VBA wires up a
double-click router so each button cell runs its corresponding macro.

Pacific Seeds brand colours throughout (white + blue, navy headers).
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Protection, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "output"
SEEDS = ROOT / "seeds"

# ---------- Pacific Seeds palette ----------
PS_NAVY     = "092A40"
PS_BLUE     = "0678CD"
PS_DEEP     = "0075BD"
PS_SKY      = "3AB3E5"
PS_LIGHT    = "8FCEEE"
PS_PALE     = "E9F3FC"
PS_WHITE    = "FFFFFF"
PS_GREY     = "F3F7FB"
PS_BORDER   = "D8E3ED"
PS_ORANGE   = "E45138"
PS_WHEAT    = "DDB318"
PS_GREEN    = "28A745"
PS_INK      = "092A40"
PS_INK_SOFT = "34526E"
PS_MUTED    = "5A7896"

THIN  = Side(style="thin",  color=PS_BORDER)
MED   = Side(style="medium", color=PS_BLUE)
HAIR  = Side(style="hair",  color=PS_BORDER)


# ---------- Style helpers ----------
def banner(ws, title: str, subtitle: str = "", cols: int = 12) -> int:
    """Pacific Seeds branded banner; returns the next free row."""
    end = get_column_letter(cols)
    ws.merge_cells(f"A1:{end}1")
    c = ws["A1"]
    c.value = "PACIFIC SEEDS"
    c.font = Font(name="Calibri", size=10, bold=True, color=PS_SKY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=PS_NAVY)
    ws.row_dimensions[1].height = 18

    ws.merge_cells(f"A2:{end}2")
    c = ws["A2"]
    c.value = title
    c.font = Font(name="Calibri", size=18, bold=True, color=PS_WHITE)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c.fill = PatternFill("solid", fgColor=PS_NAVY)
    ws.row_dimensions[2].height = 30

    if subtitle:
        ws.merge_cells(f"A3:{end}3")
        c = ws["A3"]
        c.value = subtitle
        c.font = Font(size=11, italic=True, color=PS_INK_SOFT)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.fill = PatternFill("solid", fgColor=PS_PALE)
        ws.row_dimensions[3].height = 22
        next_row = 5
    else:
        # Thin sky underline
        ws.merge_cells(f"A3:{end}3")
        c = ws["A3"]
        c.fill = PatternFill("solid", fgColor=PS_SKY)
        ws.row_dimensions[3].height = 4
        next_row = 5
    return next_row


def header_row(ws, row: int, headers: list[str], start_col: int = 1) -> None:
    for i, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(bold=True, color=PS_NAVY, size=11)
        c.fill = PatternFill("solid", fgColor=PS_PALE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=THIN, bottom=MED, left=THIN, right=THIN)
    ws.row_dimensions[row].height = 28


def button_cell(ws, cell_range: str, label: str, macro: str,
                color: str = PS_BLUE, text_color: str = PS_WHITE,
                description: str = "") -> None:
    """Style a merged cell to look like a button. Macro name stored in a comment
    AND in a parallel hidden cell for the VBA router."""
    ws.merge_cells(cell_range)
    top_left = cell_range.split(":")[0]
    c = ws[top_left]
    c.value = label
    c.font = Font(bold=True, size=13, color=text_color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor=color)
    thick = Side(style="medium", color=color)
    c.border = Border(top=thick, bottom=thick, left=thick, right=thick)
    if description:
        # Tooltip via comment
        from openpyxl.comments import Comment
        c.comment = Comment(f"{description}\n\n(Double-click to run: {macro})", "PS Workflow")


# ---------- Build: Nursery Template ----------
def build_nursery_template() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Home ----
    home = wb.create_sheet("Home")
    home.sheet_view.showGridLines = False

    next_row = banner(home, "Nursery Workflow", "Click ▶ buttons to run each step in order", cols=14)

    # Quick status box
    home.merge_cells(f"A{next_row}:E{next_row+3}")
    c = home[f"A{next_row}"]
    c.value = ("📋  This workbook follows the proposed Sorghum nursery workflow.\n"
               "    Run each step in order. Stats below auto-refresh after each step.")
    c.font = Font(size=11, color=PS_INK_SOFT)
    c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
    c.fill = PatternFill("solid", fgColor=PS_GREY)
    home.row_dimensions[next_row].height = 50
    next_row += 5

    # ---- Workflow steps (3 phases, color-coded) ----
    home.cell(row=next_row, column=1, value="PHASE 1 — PRE-FIELD PREP").font = Font(
        bold=True, color=PS_NAVY, size=12)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 1

    prep_steps = [
        ("▶ 1. Initialise from PRISM export",
         "btnInitNursery",
         "Reads the PRISM 'Nursery site' export sheet, creates all workflow tabs, "
         "and sets up the workbook structure."),
        ("▶ 2. Build Nursery list",
         "btnBuildNurseryList",
         "Picks all unique Source IDs from Nursery site, sorted A→Z. Adds Repeats "
         "and Qty Required (1.4 × Repeats). Asks for qty per packet."),
        ("▶ 3. Design Field Map",
         "btnDesignFieldMap",
         "Opens the Field Map tab; manually mark spike numbers, forward/reverse runs."),
        ("▶ 4. Generate Packet Prep & QR labels",
         "btnGeneratePacketPrep",
         "Runs the 13-step Packet Prep workflow: sort, insert Plot/Spike/Rack/QR columns, "
         "merge fields, filter rows, fill spike numbers."),
        ("▶ 5. Sort packets for racking (LSD Radix)",
         "btnSortForRacking",
         "Performs the Least-Significant-Digit Radix sort so packets stack in rack order."),
    ]

    field_steps = [
        ("▶ 6. Record replacement (Packeting/Planting)",
         "btnAddReplacement",
         "Opens a dialog to capture a replacement event with stage dropdown."),
        ("▶ 7. Record planting error",
         "btnAddPlantingError",
         "Logs a planting error in the Replacements and errors tab."),
        ("▶ 8. Spray track + date recording",
         "btnRecordSpray",
         "Adds a TFMSA / IMI / HPPD spray application with date."),
        ("▶ 9. AB bag pulling",
         "btnRecordABPull",
         "Records pulled bags with date and count."),
    ]

    post_steps = [
        ("▶ 10. Pull updated Nursery site from PRISM",
         "btnImportUpdated",
         "Imports the refreshed PRISM export into the 'Updated nursery site' tab."),
        ("▶ 11. Generate Fieldbook",
         "btnGenerateFieldbook",
         "Runs the Fieldbook template VBA: reorders columns, adds Bagging Info/Comments, "
         "serpentine sort, sets landscape print layout."),
        ("▶ 12. Refresh dashboard",
         "btnRefreshDashboard",
         "Updates the local stats panel below."),
        ("▶ 13. Push to Hub",
         "btnPushToHub",
         "Writes this nursery's summary to the shared registry.csv so the Nursery Hub "
         "workbook sees it on its dashboard."),
    ]

    def render_step_block(start_row: int, steps: list, color: str) -> int:
        r = start_row
        for label, macro, desc in steps:
            cell_range = f"A{r}:E{r+1}"
            button_cell(home, cell_range, label, macro, color=color, description=desc)
            # Macro name stored in hidden col G for the VBA router
            home.cell(row=r, column=7, value=macro).font = Font(color="BBBBBB", size=9)
            home.cell(row=r, column=7).alignment = Alignment(indent=1)
            # Description in F
            home.cell(row=r, column=6, value=desc).font = Font(
                size=10, color=PS_INK_SOFT, italic=True)
            home.cell(row=r, column=6).alignment = Alignment(
                wrap_text=True, vertical="center", indent=1)
            home.merge_cells(start_row=r, end_row=r+1, start_column=6, end_column=6)
            home.row_dimensions[r].height = 22
            home.row_dimensions[r+1].height = 22
            r += 3
        return r

    next_row = render_step_block(next_row, prep_steps, PS_BLUE)
    next_row += 1

    home.cell(row=next_row, column=1, value="PHASE 2 — FIELD OPERATIONS").font = Font(
        bold=True, color=PS_NAVY, size=12)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 1
    next_row = render_step_block(next_row, field_steps, PS_DEEP)
    next_row += 1

    home.cell(row=next_row, column=1, value="PHASE 3 — POST-FIELD & SYNC").font = Font(
        bold=True, color=PS_NAVY, size=12)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 1
    next_row = render_step_block(next_row, post_steps, PS_NAVY)
    next_row += 2

    # ---- Local dashboard ----
    home.cell(row=next_row, column=1, value="LIVE DASHBOARD").font = Font(
        bold=True, color=PS_NAVY, size=14)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 1
    home.cell(row=next_row, column=1, value="Refreshed by step 12. Pulled from the working sheets.").font = Font(
        size=10, color=PS_MUTED, italic=True)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 2

    metrics = [
        ("Total packets", "DASH_TotalPackets", "=IFERROR(COUNTA('Nursery site'!A:A)-1,0)"),
        ("Unique source IDs", "DASH_UniqueSources", "=IFERROR(COUNTA('Nursery list'!A:A)-1,0)"),
        ("Replacements logged", "DASH_Replacements",
         "=IFERROR(COUNTIF('Replacements and errors'!D:D,\"replacement\"),0)"),
        ("Planting errors logged", "DASH_PlantingErrors",
         "=IFERROR(COUNTIF('Replacements and errors'!D:D,\"planting_error\"),0)"),
        ("Spray events", "DASH_SprayEvents",
         "=IFERROR(COUNTA('Additionals'!A:A)-2,0)"),
        ("Last synced to Hub", "DASH_LastSync", "Never"),
    ]
    for i, (label, name, formula) in enumerate(metrics):
        col = (i % 3) * 4 + 1   # 1, 5, 9
        row_block = next_row + (i // 3) * 5
        # Label
        home.cell(row=row_block, column=col, value=label).font = Font(
            bold=True, color=PS_MUTED, size=10)
        home.cell(row=row_block, column=col).alignment = Alignment(indent=1)
        # Big value
        target = home.cell(row=row_block+1, column=col, value=formula)
        target.font = Font(bold=True, size=24, color=PS_BLUE)
        target.alignment = Alignment(horizontal="left", indent=1)
        home.merge_cells(start_row=row_block+1, end_row=row_block+2,
                         start_column=col, end_column=col+2)
        # Named range for VBA to reference
        try:
            dn = DefinedName(name=name,
                             attr_text=f"Home!${get_column_letter(col)}${row_block+1}")
            wb.defined_names[name] = dn
        except Exception:
            pass

    # Column widths
    for col_letter, width in [("A", 12), ("B", 24), ("C", 24), ("D", 14),
                              ("E", 18), ("F", 60), ("G", 28)]:
        home.column_dimensions[col_letter].width = width

    # ---- Settings tab ----
    settings = wb.create_sheet("Settings")
    banner(settings, "Settings", "Workbook & Hub configuration", cols=4)
    settings.cell(row=5, column=1, value="Setting").font = Font(bold=True, color=PS_NAVY)
    settings.cell(row=5, column=2, value="Value").font = Font(bold=True, color=PS_NAVY)
    settings.cell(row=5, column=3, value="Notes").font = Font(bold=True, color=PS_NAVY)
    for r, (k, v, note) in enumerate([
        ("Nursery code", "(set after step 1)", "Short code, e.g. AUGT1-26S-IMI"),
        ("Nursery name", "", "Full nursery name"),
        ("Breeder", "", "Lead breeder for this nursery"),
        ("Season", "", "e.g. 2026S"),
        ("Hub registry folder", "~/Documents/PacificSeeds/Nurseries/",
         "Shared folder where Hub looks for registry.csv. Same on every machine."),
        ("Qty per packet", 1.4, "Multiplier for required qty per Source ID"),
        ("Filter bulk treatment above", 10, "Source IDs with > X reps go to bulk list"),
    ], start=6):
        settings.cell(row=r, column=1, value=k).font = Font(color=PS_INK_SOFT)
        sv = settings.cell(row=r, column=2, value=v)
        sv.fill = PatternFill("solid", fgColor=PS_PALE)
        sv.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        settings.cell(row=r, column=3, value=note).font = Font(italic=True, color=PS_MUTED, size=10)
    settings.column_dimensions["A"].width = 26
    settings.column_dimensions["B"].width = 42
    settings.column_dimensions["C"].width = 55

    # ---- Data tabs (the 12 workflow tabs) ----
    workflow_tabs = [
        ("Nursery site", "Paste the PRISM export here (or click Step 1).",
         ["Range", "Row", "Material ID", "Source ID", "CMS reaction", "Generation",
          "Inbred Code", "Pedigree", "Hybrid Code", "Trait Name", "Plant #",
          "Loc Seq#", "SubSeq Flag", "Entry Book Project", "Entry Book Name", "Entry #"]),
        ("Field Map", "Manual grid. Mark range/row numbers and spike zones here.", []),
        ("Material Map", "Paste PRISM Material Map here.", []),
        ("Nursery data",
         "Header info — nursery name, season, breeder. Populated from Settings.", []),
        ("Nursery list", "Auto-built by Step 2 from Nursery site.",
         ["Source ID", "Repeats", "Qty Required", "Inbred Code", "Hybrid Code",
          "Treatment", "Notes"]),
        ("Packet prep", "Auto-built by Step 4.",
         ["QR payload", "Plot", "Range", "Row", "Spike", "Rack Order",
          "Material ID", "Source ID", "Generation", "CMS reaction", "Comments"]),
        ("Replacements and errors", "Log of all replacements and planting errors.",
         ["Captured At", "Tech", "Plot", "Type", "Stage", "Original Source ID",
          "Replaced With", "Severity", "Note", "Status"]),
        ("Updated nursery site", "Drop the refreshed PRISM export here for step 10.", []),
        ("Fieldbook", "Auto-built by Step 11 — print-ready.",
         ["Range", "Row", "R_R", "Crossed bags", "Bagging Info", "Material ID",
          "Source ID", "Gen", "CMS", "Comments"]),
        ("Operations", "Growth stage tracker, colour-coded.",
         ["Stage", "Plan", "Done date", "Comments"]),
        ("Comments", "Free-form team notes.", ["Date", "Tech", "Topic", "Comment"]),
        ("Additionals", "TFMSA spray, AB date recording, AB bag pulling.",
         ["Captured At", "Tech", "Plot", "Operation", "Product / Event",
          "Date", "Count", "Note"]),
    ]

    for name, subtitle, headers in workflow_tabs:
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False
        n = banner(ws, name, subtitle, cols=max(8, len(headers) or 8))
        if headers:
            header_row(ws, n, headers)
            # Width hints
            for i, _ in enumerate(headers, start=1):
                ws.column_dimensions[get_column_letter(i)].width = 16
            # Freeze under the header
            ws.freeze_panes = ws.cell(row=n + 1, column=1)
        else:
            ws.cell(row=n, column=1, value="(no auto-headers — edit freely)").font = Font(
                italic=True, color=PS_MUTED)

    # Pre-fill Operations with growth stages
    ops = wb["Operations"]
    stage_colors = [
        ("Seedling",      "FEC000"),
        ("Vegetative",    "A9D08E"),
        ("Heading",       "FF6B6B"),
        ("Flowering",     "ADD8E6"),
        ("Grain filling", "FFFF99"),
        ("Harvest",       "CCFFCC"),
        ("Post Harvest",  "CCCCFF"),
    ]
    # Operations banner ends at row 4, headers at 5
    for i, (stage, rgb) in enumerate(stage_colors, start=6):
        c = ops.cell(row=i, column=1, value=stage)
        c.font = Font(bold=True, color=PS_NAVY)
        c.fill = PatternFill("solid", fgColor=rgb)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        for col in (2, 3, 4):
            ops.cell(row=i, column=col).border = Border(
                top=THIN, bottom=THIN, left=THIN, right=THIN)
            ops.cell(row=i, column=col).alignment = Alignment(
                wrap_text=True, vertical="top")
        ops.row_dimensions[i].height = 40

    # Set up data validation on Replacements: Type column
    rep = wb["Replacements and errors"]
    dv_type = DataValidation(type="list",
        formula1='"replacement,planting_error,spray,ab_pull,note"', allow_blank=True)
    dv_type.add(f"D6:D5000")
    rep.add_data_validation(dv_type)
    dv_stage = DataValidation(type="list",
        formula1='"Packeting,Planting"', allow_blank=True)
    dv_stage.add(f"E6:E5000")
    rep.add_data_validation(dv_stage)
    dv_status = DataValidation(type="list",
        formula1='"Open,In PRISM,Closed"', allow_blank=True)
    dv_status.add(f"J6:J5000")
    rep.add_data_validation(dv_status)

    # Re-order tabs: Home first
    wb.move_sheet("Home", offset=-wb.sheetnames.index("Home"))

    OUT.mkdir(exist_ok=True)
    path = OUT / "Nursery_Template.xlsx"
    wb.save(path)
    return path


# ---------- Build: Nursery Hub ----------
def build_nursery_hub() -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    # Home
    home = wb.create_sheet("Home")
    home.sheet_view.showGridLines = False
    next_row = banner(home, "Pacific Seeds · Nursery Hub",
                      "Central view of every nursery across the program", cols=8)

    button_cell(home, f"A{next_row}:C{next_row+1}", "▶ Refresh Dashboard",
                "hubRefreshDashboard", color=PS_BLUE,
                description="Re-reads registry.csv from the shared folder and rebuilds the Dashboard tab.")
    home.row_dimensions[next_row].height = 22
    home.row_dimensions[next_row+1].height = 22

    button_cell(home, f"E{next_row}:G{next_row+1}", "▶ Register This Hub Folder",
                "hubRegisterFolder", color=PS_DEEP,
                description="Sets the shared folder where each Nursery workbook will write its summary.")

    next_row += 3
    button_cell(home, f"A{next_row}:C{next_row+1}", "▶ Open Nursery Folder",
                "hubOpenFolder", color=PS_NAVY,
                description="Opens the registry folder in Finder / File Explorer so you can pick a nursery to edit.")
    home.row_dimensions[next_row].height = 22
    home.row_dimensions[next_row+1].height = 22

    button_cell(home, f"E{next_row}:G{next_row+1}", "▶ New Nursery from Template",
                "hubCreateFromTemplate", color=PS_ORANGE,
                description="Clones Nursery_Template.xlsm into the registry folder with a new nursery code.")

    next_row += 4
    home.cell(row=next_row, column=1, value="QUICK STATS").font = Font(
        bold=True, color=PS_NAVY, size=14)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 1
    home.cell(row=next_row, column=1, value="Aggregated across every nursery registered in this Hub.").font = Font(
        size=10, color=PS_MUTED, italic=True)
    home.cell(row=next_row, column=1).alignment = Alignment(indent=1)
    next_row += 2

    hub_metrics = [
        ("Nurseries registered", "HUB_NurseryCount"),
        ("Total packets across all nurseries", "HUB_TotalPackets"),
        ("Replacements logged", "HUB_TotalReplacements"),
        ("Planting errors logged", "HUB_TotalErrors"),
    ]
    for i, (label, name) in enumerate(hub_metrics):
        col = (i % 2) * 4 + 1
        row_block = next_row + (i // 2) * 5
        home.cell(row=row_block, column=col, value=label).font = Font(
            bold=True, color=PS_MUTED, size=10)
        home.cell(row=row_block, column=col).alignment = Alignment(indent=1)
        target = home.cell(row=row_block+1, column=col, value=0)
        target.font = Font(bold=True, size=26, color=PS_BLUE)
        target.alignment = Alignment(horizontal="left", indent=1)
        home.merge_cells(start_row=row_block+1, end_row=row_block+2,
                         start_column=col, end_column=col+2)
        try:
            dn = DefinedName(name=name,
                             attr_text=f"Home!${get_column_letter(col)}${row_block+1}")
            wb.defined_names[name] = dn
        except Exception:
            pass

    for col_letter, width in [("A", 24), ("B", 24), ("C", 24), ("D", 6),
                              ("E", 24), ("F", 24), ("G", 24), ("H", 8)]:
        home.column_dimensions[col_letter].width = width

    # Dashboard
    dash = wb.create_sheet("Dashboard")
    dash.sheet_view.showGridLines = False
    n = banner(dash, "Nurseries — Live Dashboard",
               "Auto-populated from registry.csv when you press Refresh.", cols=10)
    dash_headers = ["Nursery code", "Season", "Breeder", "Packets",
                    "Replacements", "Errors", "Sprays", "Last update",
                    "File path", "Status"]
    header_row(dash, n, dash_headers)
    dash.freeze_panes = dash.cell(row=n + 1, column=1)
    for i, w in enumerate([22, 10, 16, 10, 14, 10, 10, 22, 50, 14], start=1):
        dash.column_dimensions[get_column_letter(i)].width = w

    # Settings
    settings = wb.create_sheet("Settings")
    settings.sheet_view.showGridLines = False
    banner(settings, "Hub Settings", "Set once, then Refresh.", cols=4)
    settings_rows = [
        ("Registry folder", "~/Documents/PacificSeeds/Nurseries/",
         "All nurseries write registry rows into this folder. Same on every workstation."),
        ("Registry file", "registry.csv",
         "Shared CSV inside the folder above. Don't rename — Hub looks for this name."),
        ("Template path", "~/Documents/PacificSeeds/Templates/Nursery_Template.xlsm",
         "Used by 'New Nursery from Template' button."),
    ]
    settings.cell(row=5, column=1, value="Setting").font = Font(bold=True, color=PS_NAVY)
    settings.cell(row=5, column=2, value="Value").font = Font(bold=True, color=PS_NAVY)
    settings.cell(row=5, column=3, value="Notes").font = Font(bold=True, color=PS_NAVY)
    for r, (k, v, note) in enumerate(settings_rows, start=6):
        settings.cell(row=r, column=1, value=k).font = Font(color=PS_INK_SOFT)
        sv = settings.cell(row=r, column=2, value=v)
        sv.fill = PatternFill("solid", fgColor=PS_PALE)
        sv.border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
        settings.cell(row=r, column=3, value=note).font = Font(italic=True, color=PS_MUTED, size=10)
        # Named ranges for VBA
        nm = "HUB_" + k.replace(" ", "_")
        try:
            wb.defined_names[nm] = DefinedName(name=nm,
                                               attr_text=f"Settings!$B${r}")
        except Exception:
            pass
    settings.column_dimensions["A"].width = 22
    settings.column_dimensions["B"].width = 55
    settings.column_dimensions["C"].width = 60

    wb.move_sheet("Home", offset=-wb.sheetnames.index("Home"))

    OUT.mkdir(exist_ok=True)
    path = OUT / "Nursery_Hub.xlsx"
    wb.save(path)
    return path


def _maybe_bake(xlsx_path: Path, label: str) -> Path:
    """If a seed vbaProject.bin exists for this workbook, bake it in and
    return the .xlsm path. Otherwise just return the .xlsx path unchanged.
    Seeds live in excel_workflow/seeds/<stem>.vbaProject.bin and are
    created by running extract_seed.py against a known-good .xlsm."""
    seed = SEEDS / f"{xlsx_path.stem}.vbaProject.bin"
    if not seed.exists():
        print(f"     ℹ {label}: no seed at {seed.relative_to(ROOT)} — "
              f"keeping .xlsx only. Run extract_seed.py once to lock in VBA.")
        return xlsx_path

    from bake_vba import bake_vba
    xlsm = xlsx_path.with_suffix(".xlsm")
    bake_vba(xlsx_path, seed, xlsm)
    print(f"     ✓ Baked VBA into {xlsm.name} (seed: {seed.name})")
    return xlsm


def main() -> None:
    print("→ Building workbooks…")
    t = build_nursery_template()
    print(f"  · {t.name}")
    out_t = _maybe_bake(t, "Template")

    h = build_nursery_hub()
    print(f"  · {h.name}")
    out_h = _maybe_bake(h, "Hub")

    print()
    print("Outputs:")
    print(f"  {out_t}")
    print(f"  {out_h}")


if __name__ == "__main__":
    main()
