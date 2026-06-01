"""FastAPI sync backend for the Sorghum nursery PWA.

Endpoints:
  GET  /                              → serves the PWA
  GET  /nursery/{code}/manifest       → packet list for the device's offline cache
  POST /sync                          → batch upsert of captured events (idempotent on uuid)
  GET  /nursery/{code}/events         → debugging: list synced events
"""
from __future__ import annotations

import json
import os
import sqlite3
import sys
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

ROOT = Path(__file__).resolve().parent.parent
# Per-user dirs can be overridden via env (set by the desktop launcher on Windows).
DATA_DIR = Path(os.environ.get("PS_DATA_DIR", str(ROOT / "data")))
OUT_DIR  = Path(os.environ.get("PS_OUTPUT_DIR", str(ROOT / "output")))
DATA_DIR.mkdir(parents=True, exist_ok=True)
OUT_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "nursery.sqlite"
PWA_DIR = ROOT / "pwa"
SCRIPTS_DIR = ROOT / "scripts"

# Make our own scripts importable
sys.path.insert(0, str(SCRIPTS_DIR))

app = FastAPI(title="Sorghum Nursery Sync")


def db() -> sqlite3.Connection:
    if not DB_PATH.exists():
        raise HTTPException(503, f"Database not found at {DB_PATH}. Run nursery_init.py first.")
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


class Event(BaseModel):
    uuid: str
    packet_uuid: str | None = None
    nursery_code: str
    type: str
    payload: dict[str, Any] = {}
    tech_id: str | None = None
    captured_at: str
    gps_lat: float | None = None
    gps_lon: float | None = None


class SyncRequest(BaseModel):
    events: list[Event]


class SyncResult(BaseModel):
    accepted: int
    duplicates: int


@app.get("/nursery/{code}/manifest")
def manifest(code: str) -> dict:
    conn = db()
    rows = conn.execute("""
        SELECT uuid, plot, range_n, row_n, spike, rack_order,
               material_id, source_id, inbred_code, generation, cms, comments
        FROM packets WHERE nursery_code = ?
        ORDER BY spike, rack_order
    """, (code,)).fetchall()
    if not rows:
        raise HTTPException(404, f"No packets for nursery '{code}'")
    return {
        "nursery_code": code,
        "packet_count": len(rows),
        "packets": [dict(r) for r in rows],
    }


@app.get("/nurseries")
def nurseries() -> list[str]:
    conn = db()
    return [r["code"] for r in conn.execute("SELECT code FROM nurseries ORDER BY code")]


@app.post("/sync", response_model=SyncResult)
def sync(req: SyncRequest) -> SyncResult:
    conn = db()
    accepted = 0
    duplicates = 0
    for e in req.events:
        cur = conn.execute("SELECT 1 FROM events WHERE uuid = ?", (e.uuid,)).fetchone()
        if cur:
            duplicates += 1
            continue
        conn.execute("""
            INSERT INTO events(uuid, packet_uuid, nursery_code, type, payload,
                tech_id, captured_at, gps_lat, gps_lon)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (e.uuid, e.packet_uuid, e.nursery_code, e.type,
              json.dumps(e.payload), e.tech_id, e.captured_at, e.gps_lat, e.gps_lon))
        accepted += 1
    conn.commit()
    return SyncResult(accepted=accepted, duplicates=duplicates)


@app.get("/nursery/{code}/events")
def events(code: str, limit: int = 200) -> list[dict]:
    conn = db()
    rows = conn.execute("""
        SELECT e.uuid, e.packet_uuid, e.nursery_code, e.type, e.payload,
               e.tech_id, e.captured_at, e.gps_lat, e.gps_lon, e.synced_at,
               p.plot, p.material_id, p.source_id
        FROM events e LEFT JOIN packets p ON p.uuid = e.packet_uuid
        WHERE e.nursery_code = ?
        ORDER BY e.captured_at DESC
        LIMIT ?
    """, (code, limit)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["payload"] = json.loads(d["payload"]) if d["payload"] else {}
        out.append(d)
    return out


@app.get("/nursery/{code}/summary")
def summary(code: str) -> dict:
    conn = db()
    pkt = conn.execute(
        "SELECT COUNT(*) FROM packets WHERE nursery_code = ?", (code,)
    ).fetchone()[0]
    type_counts = {r["type"]: r["c"] for r in conn.execute(
        "SELECT type, COUNT(*) c FROM events WHERE nursery_code = ? GROUP BY type", (code,)
    )}
    tech_counts = {r["tech_id"] or "(unknown)": r["c"] for r in conn.execute(
        "SELECT tech_id, COUNT(*) c FROM events WHERE nursery_code = ? GROUP BY tech_id", (code,)
    )}
    last = conn.execute(
        "SELECT MAX(captured_at) FROM events WHERE nursery_code = ?", (code,)
    ).fetchone()[0]
    return {
        "nursery_code": code,
        "packet_count": pkt,
        "event_counts_by_type": type_counts,
        "event_counts_by_tech": tech_counts,
        "last_activity": last,
    }


TYPE_LABEL = {
    "replacement": "Replacements",
    "planting_error": "Planting errors",
    "spray": "Spray applied",
    "ab_pull": "AB bags pulled",
    "date_record": "Date records",
    "note": "Notes",
}


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard() -> str:
    conn = db()
    codes = [r["code"] for r in conn.execute("SELECT code FROM nurseries ORDER BY code")]
    cards = []
    for code in codes:
        s = summary(code)
        events_html = ""
        recent = conn.execute("""
            SELECT e.type, e.tech_id, e.captured_at, e.payload, p.plot
            FROM events e LEFT JOIN packets p ON p.uuid = e.packet_uuid
            WHERE e.nursery_code = ?
            ORDER BY e.captured_at DESC LIMIT 5
        """, (code,)).fetchall()
        for r in recent:
            payload = json.loads(r["payload"]) if r["payload"] else {}
            label = TYPE_LABEL.get(r["type"], r["type"])
            detail = payload.get("replaced_with") or payload.get("error_kind") or \
                     payload.get("product") or payload.get("event_label") or \
                     payload.get("note") or ""
            events_html += (
                f"<tr><td>{r['captured_at'][:16].replace('T',' ')}</td>"
                f"<td>{r['tech_id'] or ''}</td>"
                f"<td>Plot {r['plot'] or payload.get('plot','?')}</td>"
                f"<td>{label}</td><td>{detail[:40]}</td></tr>"
            )
        if not events_html:
            events_html = '<tr><td colspan="5" class="muted">No events yet.</td></tr>'

        type_chips = " ".join(
            f'<span class="chip">{TYPE_LABEL.get(t, t)}: {n}</span>'
            for t, n in sorted(s["event_counts_by_type"].items())
        ) or '<span class="muted">no events</span>'
        tech_chips = " ".join(
            f'<span class="chip tech">{t}: {n}</span>'
            for t, n in sorted(s["event_counts_by_tech"].items())
        ) or '<span class="muted">none</span>'
        last = s["last_activity"] or "—"

        cards.append(f"""
        <section class="card">
          <header>
            <h2>{code}</h2>
            <span class="pkt">{s['packet_count']} packets</span>
          </header>
          <div class="meta"><b>Last activity:</b> {last}</div>
          <div class="meta"><b>By type:</b> {type_chips}</div>
          <div class="meta"><b>By tech:</b> {tech_chips}</div>
          <table>
            <thead><tr><th>When</th><th>Tech</th><th>Plot</th><th>Type</th><th>Detail</th></tr></thead>
            <tbody>{events_html}</tbody>
          </table>
          <div class="links">
            <a href="/nursery/{code}/manifest">manifest json</a> ·
            <a href="/nursery/{code}/events">events json</a> ·
            <a href="/nursery/{code}/summary">summary json</a>
          </div>
        </section>
        """)

    if not cards:
        cards = ['<section class="card empty">No nurseries ingested. Run <code>nursery_init.py</code>.</section>']

    return f"""<!doctype html><html><head><meta charset="utf-8">
    <title>Pacific Seeds · Nursery Dashboard</title>
    <link rel="icon" href="/static/ps-logo.svg" type="image/svg+xml" />
    <style>
      :root {{ --ps-blue:#0678CD; --ps-deep:#0075BD; --ps-navy:#092A40;
              --ps-sky:#3AB3E5; --ps-light:#8FCEEE;
              --ps-orange:#e45138; --ps-wheat:#ddb318; --ps-green:#28a745;
              --ink:#0f2540; --muted:#5a7896; }}
      * {{ box-sizing: border-box; }}
      body {{ font-family: -apple-system, system-ui, "Segoe UI", Roboto, sans-serif;
             background:#f3f7fb; margin:0; color:var(--ink); }}
      header.top {{ background:var(--ps-navy); color:#fff; padding:14px 24px;
                   display:flex; align-items:center; gap:16px;
                   border-bottom:3px solid var(--ps-sky); }}
      header.top .logo {{ height:32px; }}
      header.top .titles {{ display:flex; flex-direction:column; line-height:1.1; }}
      header.top .titles .sub {{ font-size:11px; color:var(--ps-sky);
                                letter-spacing:1.5px; text-transform:uppercase; font-weight:600; }}
      header.top .titles .main {{ font-size:16px; font-weight:600; }}
      header.top .grow {{ flex:1; }}
      header.top a {{ color:var(--ps-sky); text-decoration:none; font-size:14px; font-weight:500; }}
      header.top a:hover {{ color:#fff; }}
      main {{ max-width: 1080px; margin: 28px auto; padding: 0 20px; }}
      .card {{ background:#fff; border:1px solid #dde6ef; border-radius:12px; padding:20px;
              margin-bottom:20px; box-shadow:0 1px 3px rgba(9,42,64,0.06); }}
      .card header {{ display:flex; align-items:center; gap:12px; margin-bottom:8px;
                     padding-bottom:10px; border-bottom:1px solid #eef2f6; }}
      .card h2 {{ margin:0; font-size:18px; color:var(--ps-navy); font-weight:600; }}
      .pkt {{ background:var(--ps-blue); color:#fff; padding:4px 12px; border-radius:999px;
             font-size:12px; font-weight:600; letter-spacing:0.3px; }}
      .meta {{ font-size:13px; color:#34526e; margin: 6px 0; }}
      .meta b {{ color:var(--ps-navy); }}
      .chip {{ display:inline-block; background:#e9f3fc; color:var(--ps-deep);
              padding:3px 10px; border-radius:999px; font-size:12px; margin-right:4px;
              font-weight:500; }}
      .chip.tech {{ background:#fdeee6; color:#b54724; }}
      .muted {{ color:var(--muted); font-style:italic; }}
      table {{ width:100%; border-collapse:collapse; margin-top:12px; font-size:13px; }}
      th, td {{ text-align:left; padding:8px 10px; border-bottom:1px solid #eef2f6; }}
      th {{ background:#f7fafd; font-weight:600; color:var(--ps-navy);
            text-transform:uppercase; font-size:11px; letter-spacing:0.5px; }}
      tbody tr:hover {{ background:#fafcfe; }}
      .links {{ margin-top:12px; font-size:12px; }}
      .links a {{ color:var(--ps-blue); text-decoration:none; margin-right:6px; }}
      .links a:hover {{ text-decoration:underline; }}
      .empty {{ text-align:center; color:var(--muted); padding:50px; }}
      .empty code {{ background:#eef2f6; padding:2px 6px; border-radius:4px;
                    color:var(--ps-navy); font-size:13px; }}
    </style></head><body>
    <header class="top">
      <img src="/static/ps-logo-white.svg" alt="Pacific Seeds" class="logo" />
      <div class="titles">
        <span class="sub">R&amp;D Sorghum</span>
        <span class="main">Nursery Dashboard</span>
      </div>
      <span class="grow"></span>
      <a href="/">Open field tech app →</a>
    </header>
    <main>{''.join(cards)}</main>
    </body></html>"""


#============================================================================
#  Admin (Phase 1 + post-field) endpoints — Excel-workflow parity
#============================================================================

@app.post("/admin/inspect")
async def admin_inspect(file: UploadFile = File(...)) -> dict:
    """Quick peek into an uploaded .xlsx: returns sheet names + best guess at
    the PRISM data sheet (one with a Range/Row/Source ID header)."""
    from openpyxl import load_workbook
    tmp = Path(tempfile.gettempdir()) / f"inspect_{file.filename}"
    tmp.write_bytes(await file.read())
    try:
        wb = load_workbook(tmp, read_only=True, data_only=True)
        sheets = []
        guess = None
        for name in wb.sheetnames:
            ws = wb[name]
            # Try header row at common positions
            headers: list[str] = []
            for r in (1, 5):
                row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True), None)
                if row:
                    headers = [str(v).strip() if v is not None else "" for v in row]
                    if any(h for h in headers):
                        break
            has_prism = any(h.lower() == "range" for h in headers) and \
                        any(h.lower() == "row" for h in headers) and \
                        any(h.lower() == "source id" for h in headers)
            info = {
                "name": name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "has_prism_headers": has_prism,
                "headers_preview": [h for h in headers if h][:6],
            }
            sheets.append(info)
            if has_prism and guess is None:
                guess = name
        return {"sheets": sheets, "suggested": guess or (sheets[0]["name"] if sheets else None)}
    except Exception as e:
        raise HTTPException(400, f"Could not read file: {e}")
    finally:
        try: tmp.unlink()
        except Exception: pass


@app.post("/admin/init")
async def admin_init(
    nursery_code: str = Form(...),
    sheet: str = Form("Sheet1"),
    file: UploadFile = File(...),
) -> dict:
    """Step 1 — initialise a nursery from an uploaded PRISM export."""
    if not nursery_code.strip():
        raise HTTPException(400, "Nursery code required")
    # Save the uploaded file to a temp path, then call the script function.
    tmp = Path(tempfile.gettempdir()) / f"prism_{nursery_code}.xlsx"
    tmp.write_bytes(await file.read())
    try:
        from nursery_init import init_nursery
        result = init_nursery(tmp, nursery_code, sheet)
    except Exception as e:
        raise HTTPException(500, f"Init failed: {e}")
    finally:
        try: tmp.unlink()
        except Exception: pass
    return result


@app.get("/nursery/{code}/list")
def nursery_list(code: str) -> dict:
    """Step 2 — auto-built Nursery list: unique Source IDs with reps and qty."""
    conn = db()
    rows = conn.execute("""
        SELECT source_id, COUNT(*) AS reps,
               MAX(inbred_code) AS inbred,
               MAX(hybrid_code) AS hybrid
        FROM packets
        WHERE nursery_code = ? AND source_id IS NOT NULL AND source_id != ''
        GROUP BY source_id
        ORDER BY source_id
    """, (code,)).fetchall()
    QTY_PER = 1.4
    BULK_ABOVE = 10
    items = []
    for r in rows:
        reps = r["reps"]
        items.append({
            "source_id": r["source_id"],
            "reps": reps,
            "qty_required": round(reps * QTY_PER, 1),
            "inbred_code": r["inbred"] or "",
            "hybrid_code": r["hybrid"] or "",
            "bulk": reps > BULK_ABOVE,
        })
    return {"nursery_code": code, "qty_per_packet": QTY_PER,
            "bulk_threshold": BULK_ABOVE, "items": items}


@app.get("/nursery/{code}/workbook.xlsx")
def packets_workbook(code: str) -> FileResponse:
    """Step 4 — download the full multi-tab workbook (Packet Prep + all tabs).

    The Packet Prep tab carries the QR CODE text payload column, ready to be
    fed into a barcode printer machine on a different workstation."""
    wb = OUT_DIR / f"{code}_workbook.xlsx"
    if not wb.exists():
        raise HTTPException(404, f"Workbook not built yet for {code}. Re-run init.")
    return FileResponse(
        wb,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{code}_workbook.xlsx",
    )


@app.post("/admin/fieldbook/{code}")
def admin_fieldbook(code: str) -> FileResponse:
    """Step 11 — generate the multi-tab fieldbook xlsx for PRISM upload."""
    from fieldbook_export import main as fieldbook_main
    out = OUT_DIR / f"{code}_fieldbook.xlsx"
    OUT_DIR.mkdir(exist_ok=True)
    saved_argv = sys.argv
    sys.argv = ["fieldbook_export", "--nursery-code", code, "--out", str(out)]
    try:
        fieldbook_main()
    finally:
        sys.argv = saved_argv
    if not out.exists():
        raise HTTPException(500, "Fieldbook generation failed")
    return FileResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{code}_fieldbook.xlsx",
    )


#============================================================================
#  Field Map editor endpoints — spike layout
#============================================================================

def _ensure_field_maps_table(conn: sqlite3.Connection) -> None:
    conn.execute("""
        CREATE TABLE IF NOT EXISTS field_maps (
            nursery_code TEXT PRIMARY KEY,
            layout TEXT NOT NULL,
            updated_at TEXT DEFAULT CURRENT_TIMESTAMP
        )
    """)


@app.get("/nursery/{code}/map")
def get_field_map(code: str) -> dict:
    """Return the current spike layout — either user-saved, or derived from packets."""
    conn = db()
    _ensure_field_maps_table(conn)
    row = conn.execute(
        "SELECT layout, updated_at FROM field_maps WHERE nursery_code = ?",
        (code,),
    ).fetchone()
    if row:
        return {
            "nursery_code": code,
            "source": "saved",
            "updated_at": row["updated_at"],
            "spikes": json.loads(row["layout"]),
        }
    # Derive from current packets table.
    spikes: dict[int, list[int]] = {}
    for r in conn.execute("""
        SELECT spike, row_n
        FROM packets
        WHERE nursery_code = ? AND spike IS NOT NULL
        GROUP BY spike, row_n
        ORDER BY spike, row_n
    """, (code,)):
        spikes.setdefault(r["spike"], []).append(r["row_n"])
    return {
        "nursery_code": code,
        "source": "derived",
        "updated_at": None,
        "spikes": [
            {"spike": s, "rows": rs, "row_min": min(rs), "row_max": max(rs)}
            for s, rs in sorted(spikes.items())
        ],
    }


class FieldMapPayload(BaseModel):
    spikes: list[dict[str, Any]]   # each: {spike: int, row_min: int, row_max: int}


@app.post("/nursery/{code}/map")
def save_field_map(code: str, payload: FieldMapPayload) -> dict:
    """Save a new spike layout and re-assign every packet's spike + rack_order."""
    if not payload.spikes:
        raise HTTPException(400, "At least one spike required")
    conn = db()
    _ensure_field_maps_table(conn)

    # Validate + expand into row→spike map.
    row_to_spike: dict[int, int] = {}
    row_order_in_spike: dict[int, list[int]] = {}
    for s in payload.spikes:
        try:
            spike_no = int(s["spike"])
            rmin = int(s["row_min"])
            rmax = int(s["row_max"])
        except (KeyError, ValueError, TypeError):
            raise HTTPException(400, "Each spike needs {spike, row_min, row_max} as integers")
        if rmin > rmax:
            raise HTTPException(400, f"Spike {spike_no}: row_min ({rmin}) > row_max ({rmax})")
        rows = list(range(rmin, rmax + 1))
        row_order_in_spike[spike_no] = rows
        for r in rows:
            if r in row_to_spike:
                raise HTTPException(400, f"Row {r} assigned to multiple spikes")
            row_to_spike[r] = spike_no

    # Re-assign every packet for this nursery.
    pkts = conn.execute(
        "SELECT uuid, range_n, row_n FROM packets WHERE nursery_code = ?",
        (code,),
    ).fetchall()
    by_spike: dict[int, list[sqlite3.Row]] = {}
    for p in pkts:
        spike = row_to_spike.get(p["row_n"], 0)
        by_spike.setdefault(spike, []).append(p)
    updates = []
    for spike, items in by_spike.items():
        order = row_order_in_spike.get(spike, [])
        pos = {r: i for i, r in enumerate(order)}

        def k(pp):
            i = pos.get(pp["row_n"], 9999)
            return (pp["range_n"], -i if pp["range_n"] % 2 == 0 else i)

        items.sort(key=k)
        for i, p in enumerate(items, start=1):
            updates.append((spike, i, p["uuid"]))
    conn.executemany(
        "UPDATE packets SET spike = ?, rack_order = ? WHERE uuid = ?",
        updates,
    )
    conn.execute(
        "INSERT OR REPLACE INTO field_maps(nursery_code, layout, updated_at) VALUES (?, ?, ?)",
        (code, json.dumps([s for s in payload.spikes]), datetime.now().isoformat(sep=" ", timespec="seconds")),
    )
    conn.commit()

    return {
        "nursery_code": code,
        "spike_count": len(by_spike),
        "packets_reassigned": len(updates),
        "off_map": len(by_spike.get(0, [])),
    }


# Root = explanatory "How it works" page; PWA workflow lives at /app.
@app.get("/")
def landing() -> FileResponse:
    return FileResponse(PWA_DIR / "how-it-works.html")


@app.get("/app")
def app_index() -> FileResponse:
    return FileResponse(PWA_DIR / "index.html")


# Keep the older URL working as an alias.
@app.get("/how-it-works")
def how_it_works() -> FileResponse:
    return FileResponse(PWA_DIR / "how-it-works.html")


app.mount("/static", StaticFiles(directory=PWA_DIR), name="static")
