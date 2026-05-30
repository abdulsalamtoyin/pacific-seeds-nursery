"""Export captured events back into a multi-tab .xlsx for PRISM upload.

One sheet per event type, created only if at least one event of that type exists.
Sheets are styled with bold green headers, matching the existing Replacements tab.
"""
from __future__ import annotations

import argparse
import json
import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
DB_PATH = ROOT / "data" / "nursery.sqlite"

HEADER_FILL = PatternFill("solid", fgColor="A9D08E")
HEADER_FONT = Font(bold=True)
HEADER_ALIGN = Alignment(horizontal="center")


# Per-type sheet definition: (sheet_name, [(header, payload_key | special), ...]).
# A special key starting with "$" pulls from the event/packet row instead of payload.
SHEET_DEFS = {
    "replacement": ("Replacements done", [
        ("Plot",                "$plot"),
        ("Original Source ID",  "original_source"),
        ("Replaced with",       "replaced_with"),
        ("Stage",               "stage"),
        ("Tech",                "$tech"),
        ("Captured at",         "$captured_at"),
        ("Note",                "note"),
        ("Event UUID",          "$uuid"),
    ]),
    "planting_error": ("Planting errors", [
        ("Plot",         "$plot"),
        ("Material ID",  "$material_id"),
        ("Source ID",    "$source_id"),
        ("Error kind",   "error_kind"),
        ("Severity",     "severity"),
        ("Tech",         "$tech"),
        ("Captured at",  "$captured_at"),
        ("Note",         "note"),
        ("Event UUID",   "$uuid"),
    ]),
    "spray": ("Spray log", [
        ("Plot",         "$plot"),
        ("Product",      "product"),
        ("Date applied", "applied_on"),
        ("Rate",         "rate"),
        ("Tech",         "$tech"),
        ("Captured at",  "$captured_at"),
        ("Note",         "note"),
        ("Event UUID",   "$uuid"),
    ]),
    "ab_pull": ("AB bag pulling", [
        ("Plot",         "$plot"),
        ("Bags pulled",  "bag_count"),
        ("Date pulled",  "pulled_on"),
        ("Tech",         "$tech"),
        ("Captured at",  "$captured_at"),
        ("Note",         "note"),
        ("Event UUID",   "$uuid"),
    ]),
    "date_record": ("Date recording", [
        ("Plot",         "$plot"),
        ("Material ID",  "$material_id"),
        ("Event",        "event_label"),
        ("Date",         "occurred_on"),
        ("Tech",         "$tech"),
        ("Captured at",  "$captured_at"),
        ("Note",         "note"),
        ("Event UUID",   "$uuid"),
    ]),
    "note": ("Notes", [
        ("Plot",         "$plot"),
        ("Material ID",  "$material_id"),
        ("Note",         "note"),
        ("Tech",         "$tech"),
        ("Captured at",  "$captured_at"),
        ("Event UUID",   "$uuid"),
    ]),
}


def resolve(key: str, ev_row: sqlite3.Row, payload: dict) -> object:
    if not key.startswith("$"):
        return payload.get(key)
    field = key[1:]
    if field == "plot":
        return ev_row["plot"] or payload.get("plot")
    if field == "tech":
        return ev_row["tech_id"]
    if field == "captured_at":
        return ev_row["captured_at"]
    if field == "uuid":
        return ev_row["uuid"]
    return ev_row[field] if field in ev_row.keys() else None


def style_header(ws) -> None:
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--nursery-code", required=True)
    ap.add_argument("--out", required=True, type=Path)
    args = ap.parse_args()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT e.uuid, e.type, e.payload, e.tech_id, e.captured_at, e.packet_uuid,
               p.plot, p.material_id, p.source_id, p.spike, p.rack_order
        FROM events e LEFT JOIN packets p ON p.uuid = e.packet_uuid
        WHERE e.nursery_code = ?
        ORDER BY e.captured_at
    """, (args.nursery_code,)).fetchall()

    # Bucket by type.
    by_type: dict[str, list[sqlite3.Row]] = {}
    for r in rows:
        by_type.setdefault(r["type"], []).append(r)

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet always exists.
    summary = wb.create_sheet("Summary")
    summary.append(["Nursery", args.nursery_code])
    summary.append(["Generated rows", len(rows)])
    summary.append([])
    summary.append(["Event type", "Count"])
    style_header(summary)  # styles row 1; that's fine
    for t in sorted(by_type):
        summary.append([SHEET_DEFS.get(t, (t,))[0], len(by_type[t])])

    written = []
    for type_key, (sheet_name, cols) in SHEET_DEFS.items():
        evs = by_type.get(type_key, [])
        if not evs:
            continue
        ws = wb.create_sheet(sheet_name)
        ws.append([h for h, _ in cols])
        style_header(ws)
        for r in evs:
            payload = json.loads(r["payload"]) if r["payload"] else {}
            ws.append([resolve(k, r, payload) for _, k in cols])
        # Width tuning: longish text columns get more room.
        for i, (h, _) in enumerate(cols, start=1):
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = (
                30 if any(w in h.lower() for w in ("source", "note", "replaced", "uuid")) else 14
            )
        written.append((sheet_name, len(evs)))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    for name, n in written:
        print(f"  · {name}: {n} row(s)")
    if not written:
        print("  (no events recorded yet)")


if __name__ == "__main__":
    main()
