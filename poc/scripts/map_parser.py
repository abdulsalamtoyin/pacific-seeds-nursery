"""Parse a hand-designed Field Map sheet to recover spike assignment.

The Map is a 2D layout where:
  * one row near the top contains **field-row** numbers as column headers
    (matching the "Row" column in the PRISM export)
  * one column on the side contains **range** numbers, counting down
  * groups of consecutive header values separated by blank columns are *spikes*
  * cells contain zone labels or material check names (not packet data)

We extract: a dict { field_row -> spike_number }, plus the column-walking
order of rows inside each spike (used to drive serpentine rack-order).
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class MapLayout:
    sheet_name: str
    row_to_spike: dict[int, int]                 # field-row number -> spike number
    row_order_in_spike: dict[int, list[int]]     # spike -> field rows in column order
    header_row: int
    row_header_cols: dict[int, int]              # field row -> Map column index

    @property
    def spike_count(self) -> int:
        return len(self.row_order_in_spike)


def _is_int(v) -> bool:
    if isinstance(v, int):
        return True
    if isinstance(v, float) and v.is_integer():
        return True
    if isinstance(v, str) and v.strip().lstrip("-").isdigit():
        return True
    return False


def _to_int(v) -> int:
    if isinstance(v, str):
        return int(v.strip())
    return int(v)


def _find_header_row(rows: list[tuple]) -> int | None:
    """Find the row most likely to be the row-number header.

    Heuristic: the row with the longest run of consecutive small integers
    (1, 2, 3, …), allowing one or more blank gaps in between (spike separators).
    """
    best_row, best_score = None, 0
    for r_idx, row in enumerate(rows[:15]):  # only look near the top
        ints = [(c_idx, _to_int(v)) for c_idx, v in enumerate(row) if _is_int(v)]
        if len(ints) < 4:
            continue
        # Check that values look like 1..N (start at 1, mostly consecutive).
        values = [v for _, v in ints]
        if min(values) != 1:
            # Some maps may start at 1 in two places; allow if 1 is present.
            if 1 not in values:
                continue
        # Score = number of consecutive-pair matches (v[i+1] == v[i]+1).
        score = sum(1 for a, b in zip(values, values[1:]) if b == a + 1)
        if score > best_score:
            best_score, best_row = score, r_idx
    return best_row


def _segment_into_spikes(range_cols: list[tuple[int, int]]) -> list[list[int]]:
    """Given [(col_idx, range_n), ...] in column order, split into spikes.

    A spike boundary occurs where the column index jumps by more than 1
    (i.e. one or more blank separator columns between range numbers).
    """
    if not range_cols:
        return []
    spikes: list[list[int]] = [[]]
    prev_col = None
    for col, rng in range_cols:
        if prev_col is not None and col - prev_col > 1:
            spikes.append([])
        spikes[-1].append(rng)
        prev_col = col
    return spikes


def parse_map(xlsx_path: Path, sheet_name: str | None = None) -> MapLayout | None:
    """Return a MapLayout if a parsable Map sheet is found, else None."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    candidates = []
    if sheet_name:
        if sheet_name in wb.sheetnames:
            candidates = [sheet_name]
    else:
        candidates = [s for s in wb.sheetnames
                      if s.lower() in {"map", "field map", "fieldmap"}]
    for name in candidates:
        ws = wb[name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        hr = _find_header_row(rows)
        if hr is None:
            continue
        header = rows[hr]
        header_cells = [(c, _to_int(v)) for c, v in enumerate(header) if _is_int(v)]
        header_cells = [(c, fr) for c, fr in header_cells if 1 <= fr <= 200]
        if len(header_cells) < 4:
            continue
        spikes = _segment_into_spikes(header_cells)
        row_to_spike: dict[int, int] = {}
        order: dict[int, list[int]] = {}
        for spike_idx, field_rows in enumerate(spikes, start=1):
            order[spike_idx] = field_rows
            for fr in field_rows:
                row_to_spike[fr] = spike_idx
        return MapLayout(
            sheet_name=name,
            row_to_spike=row_to_spike,
            row_order_in_spike=order,
            header_row=hr + 1,
            row_header_cols={fr: c for c, fr in header_cells},
        )
    return None


if __name__ == "__main__":
    import sys
    p = Path(sys.argv[1])
    layout = parse_map(p)
    if not layout:
        print("No Map found.")
        sys.exit(1)
    print(f"Map sheet: {layout.sheet_name}")
    print(f"Header row: {layout.header_row}")
    print(f"Spikes: {layout.spike_count}")
    for spike, field_rows in layout.row_order_in_spike.items():
        print(f"  Spike {spike}: field rows {field_rows}")
