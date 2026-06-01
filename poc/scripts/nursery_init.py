"""Initialise a nursery from a PRISM export.

Replaces three VBA modules:
  - Recurrent file template.bas  → creates workbook tabs
  - Packetprinting.bas           → sorts, inserts Plot/SPIKE#/RACK ORDER columns
  - Fieldbook template.bas       → Fieldbook tab

Outputs:
  - data/nursery.sqlite (packets table seeded for this nursery)
  - output/<nursery>_workbook.xlsx — the tab-structured workbook, with every
    tab from the sample files populated:
      Nursery site · Map · Nursery data · Packet Prep · Nursery list ·
      Fieldbook · Replacements done · Planting error noted · Operations

The Packet Prep tab is the key output: its QR CODE column contains a
comma-joined text payload that another machine uses to print physical
barcodes. We no longer generate a printable PDF of QR labels here.
"""
from __future__ import annotations

import argparse
import sqlite3
import uuid
from collections import Counter
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from map_parser import MapLayout, parse_map

import os

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("PS_DATA_DIR", str(ROOT / "data")))
OUT_DIR  = Path(os.environ.get("PS_OUTPUT_DIR", str(ROOT / "output")))
DB_PATH = DATA_DIR / "nursery.sqlite"

# Raw PRISM export columns (in the order they appear in the sample files).
PRISM_COLS = [
    "Range", "Row", "Material ID", "Inbred Code", "Source ID", "CMS reaction",
    "Generation", "Comments", "Pedigree", "Hybrid Code", "Trait Name",
    "Plant #", "Loc Seq#", "SubSeq Flag", "Entry Book Project",
    "Entry Book Name", "Entry #",
]

# Packet Prep tab columns (25 cols, matches the AUGT1-26S-IMI sample exactly).
PACKET_PREP_COLS = [
    "QR CODE", "Range", "Row", "Plot", "SPIKE#", "RACK ORDER",
    "Thousands/Black", "Hundreds/Red", "Tens/Green", "Ones/Blue",
    "Material ID", "Inbred Code", "Source ID", "CMS reaction",
    "Generation", "Comments", "Pedigree", "Hybrid Code",
    "Trait Name", "Plant #", "Loc Seq#", "SubSeq Flag",
    "Entry Book Project", "Entry Book Name", "Entry #",
]

# Colored fonts for the four digit columns of RACK ORDER.
DIGIT_COLORS = {
    "Thousands/Black": "000000",
    "Hundreds/Red":    "C00000",
    "Tens/Green":      "00B050",
    "Ones/Blue":       "0070C0",
}

# Fieldbook tab columns (10 cols, matches the sample).
FIELDBOOK_COLS = [
    "Range", "Row", "R_R", "Crossed bags", "Bagging Info",
    "Material ID", "Source ID", "Gen", "CMS", "Comments",
]

GROWTH_STAGES = [
    ("Seedling",      (254, 192, 0)),
    ("Vegetative",    (169, 208, 142)),
    ("Heading",       (255, 0, 0)),
    ("Flowering",     (173, 216, 230)),
    ("Grain filling", (255, 255, 153)),
    ("Harvest",       (204, 255, 204)),
    ("Post Harvest",  (204, 204, 255)),
]

# Filter helpers for the specialised tabs
def _is_bc_generation(gen) -> bool:
    """BC0, BC1, BC2, BC3 etc — used for BC labels + TFMSA Spray plots."""
    if gen is None:
        return False
    s = str(gen).strip().upper()
    return s.startswith("BC")

def _is_recurrent(gen) -> bool:
    """BC* and Fn — recurrent inbred work (Date recording / Pulling bags).
    Excludes F1 (hybrids) and numbered Fn like F8 (selections)."""
    if gen is None:
        return False
    s = str(gen).strip()
    s_up = s.upper()
    if s_up.startswith("BC"):
        return True
    # Match exactly "Fn" (the lowercase generic placeholder) but not F1/F2/.../F8
    return s == "Fn" or s_up == "FN"

def _is_hybrid_f1(gen) -> bool:
    """F1 hybrid for height tracking."""
    if gen is None:
        return False
    return str(gen).strip().upper() == "F1"


def init_db() -> sqlite3.Connection:
    DATA_DIR.mkdir(exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS nurseries (
      code TEXT PRIMARY KEY,
      created_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE TABLE IF NOT EXISTS packets (
      uuid TEXT PRIMARY KEY,
      nursery_code TEXT NOT NULL REFERENCES nurseries(code),
      range_n INTEGER, row_n INTEGER, plot TEXT,
      spike INTEGER, rack_order INTEGER,
      material_id TEXT, source_id TEXT, inbred_code TEXT,
      generation TEXT, cms TEXT, comments TEXT,
      pedigree TEXT, hybrid_code TEXT,
      entry_no INTEGER
    );
    CREATE INDEX IF NOT EXISTS idx_packets_nursery ON packets(nursery_code);
    CREATE TABLE IF NOT EXISTS events (
      uuid TEXT PRIMARY KEY,
      packet_uuid TEXT REFERENCES packets(uuid),
      nursery_code TEXT NOT NULL,
      type TEXT NOT NULL,           -- replacement | planting_error | spray | ab_pull | note
      payload TEXT,                 -- JSON
      tech_id TEXT,
      captured_at TEXT NOT NULL,    -- ISO from device
      gps_lat REAL, gps_lon REAL,
      synced_at TEXT DEFAULT CURRENT_TIMESTAMP
    );
    CREATE INDEX IF NOT EXISTS idx_events_nursery ON events(nursery_code);
    """)
    return conn


def read_prism_export(path: Path, sheet: str) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    missing = [c for c in PRISM_COLS if c not in header]
    if missing:
        print(f"WARNING: PRISM export is missing columns: {missing}")
    idx = {c: header.index(c) for c in PRISM_COLS if c in header}

    def cell(r, c):
        return r[idx[c]] if c in idx and idx[c] < len(r) else None

    out = []
    for r in rows:
        if r is None or all(v is None for v in r):
            continue
        if cell(r, "Range") is None or cell(r, "Row") is None:
            continue
        out.append({
            "range_n":     int(cell(r, "Range")),
            "row_n":       int(cell(r, "Row")),
            "material_id": cell(r, "Material ID"),
            "inbred_code": cell(r, "Inbred Code"),
            "source_id":   cell(r, "Source ID"),
            "cms":         cell(r, "CMS reaction"),
            "generation":  cell(r, "Generation"),
            "comments":    cell(r, "Comments"),
            "pedigree":    cell(r, "Pedigree"),
            "hybrid_code": cell(r, "Hybrid Code"),
            "trait_name":  cell(r, "Trait Name"),
            "plant_no":    cell(r, "Plant #"),
            "loc_seq":     cell(r, "Loc Seq#"),
            "subseq_flag": cell(r, "SubSeq Flag"),
            "entry_book_project": cell(r, "Entry Book Project"),
            "entry_book_name":    cell(r, "Entry Book Name"),
            "entry_no":    cell(r, "Entry #"),
        })
    return out


def assign_spike_rack(packets: list[dict], layout: MapLayout | None) -> None:
    """Set spike + rack_order on each packet.

    If a parsed MapLayout is available, spike numbers come from the Map (1, 2, …)
    and rack order is a serpentine walk down the ranges within that spike.
    Otherwise we fall back to "spike = range_n" with per-range serpentine.
    """
    for p in packets:
        p["plot"] = f"{p['range_n']}_{p['row_n']}"

    if layout is None:
        by_range: dict[int, list[dict]] = {}
        for p in packets:
            by_range.setdefault(p["range_n"], []).append(p)
        for rng, items in by_range.items():
            reverse = (rng % 2 == 0)
            items.sort(key=lambda p: p["row_n"], reverse=reverse)
            for i, p in enumerate(items, start=1):
                p["spike"] = rng
                p["rack_order"] = i
        return

    # Group by spike using field-row → spike mapping from the Map.
    # Rack-order = serpentine walk: for each range (ascending), pick that
    # range's packets within the spike's field-rows in alternating direction.
    by_spike: dict[int, list[dict]] = {}
    for p in packets:
        spike = layout.row_to_spike.get(p["row_n"], 0)  # 0 = off-map ("Other")
        p["spike"] = spike
        by_spike.setdefault(spike, []).append(p)

    for spike, items in by_spike.items():
        field_row_order = layout.row_order_in_spike.get(spike, [])
        row_pos = {fr: i for i, fr in enumerate(field_row_order)}

        def sort_key(pp):
            rng = pp["range_n"]
            i = row_pos.get(pp["row_n"], 9999)
            # Even-numbered ranges traverse the spike in reverse (serpentine).
            return (rng, -i if rng % 2 == 0 else i)

        items.sort(key=sort_key)
        for i, p in enumerate(items, start=1):
            p["rack_order"] = i


def insert_packets(conn: sqlite3.Connection, nursery_code: str, packets: list[dict]) -> None:
    conn.execute("INSERT OR REPLACE INTO nurseries(code) VALUES (?)", (nursery_code,))
    conn.execute("DELETE FROM packets WHERE nursery_code = ?", (nursery_code,))
    for p in packets:
        p["uuid"] = uuid.uuid4().hex[:8]
        conn.execute("""
            INSERT INTO packets(uuid, nursery_code, range_n, row_n, plot, spike, rack_order,
                material_id, source_id, inbred_code, generation, cms, comments,
                pedigree, hybrid_code, entry_no)
            VALUES (:uuid, :nc, :range_n, :row_n, :plot, :spike, :rack_order,
                :material_id, :source_id, :inbred_code, :generation, :cms, :comments,
                :pedigree, :hybrid_code, :entry_no)
        """, {**p, "nc": nursery_code})
    conn.commit()


def qr_text(p: dict) -> str:
    """QR CODE payload in the original sample format:
    `Plot,Material ID,Inbred Code,Source ID,CMS,Generation,Comments`
    Empty fields keep their comma slot (so the format is parse-stable)."""
    def f(k):
        v = p.get(k)
        return "" if v is None else str(v)
    return ",".join([
        f("plot"),
        f("material_id"),
        f("inbred_code"),
        f("source_id"),
        f("cms"),
        f("generation"),
        f("comments"),
    ])


def rack_digits(rack_order: int) -> tuple[str, str, str, str]:
    """Split RACK ORDER into 4 digits, left-padded with zeros.
    Used by the color-coded packet labels — Thousands/Hundreds/Tens/Ones."""
    s = f"{int(rack_order):04d}"
    return s[0], s[1], s[2], s[3]


def _format_header_row(ws, row_idx: int = 1, navy: str = "092A40",
                       sky: str = "E9F3FC") -> None:
    """Bold header row with PS navy text on light-blue fill + bottom border."""
    bottom = Border(bottom=Side(style="medium", color=navy))
    for c in ws[row_idx]:
        c.font = Font(bold=True, color=navy, size=11)
        c.fill = PatternFill("solid", fgColor=sky)
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = bottom
    ws.row_dimensions[row_idx].height = 26


def _autosize(ws, max_cols: int, default: int = 13) -> None:
    for i in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = default


def _write_grid_map(ws, packets: list[dict], value_key: str, title: str,
                    layout: MapLayout | None = None) -> None:
    """Build a 2D grid visualisation:
       Y-axis = ranges descending, X-axis = field rows ascending,
       cells = packet[value_key] at that (range, row) position.
    Used for both Field Map (value_key='hybrid_code') and Material Map
    (value_key='material_id')."""
    if not packets:
        return
    max_range = max(p["range_n"] for p in packets)
    max_row   = max(p["row_n"]   for p in packets)

    navy = "092A40"; sky = "E9F3FC"; border_grey = "D8E3ED"
    thin = Side(style="thin", color=border_grey)
    cell_border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Title + subtitle
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color=navy)
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=min(max_row + 2, 10))
    ws.cell(row=2, column=1, value="Range numbers down the side · Field rows across the top").font = \
        Font(italic=True, color="5A7896", size=10)
    ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=min(max_row + 2, 10))

    # Header row (field row numbers)
    HDR_ROW = 3
    c = ws.cell(row=HDR_ROW, column=1, value="Rng \\ Row")
    c.font = Font(bold=True, color=navy)
    c.fill = PatternFill("solid", fgColor=sky)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = cell_border
    for col_idx, fr in enumerate(range(1, max_row + 1), start=2):
        c = ws.cell(row=HDR_ROW, column=col_idx, value=fr)
        c.font = Font(bold=True, color=navy)
        c.fill = PatternFill("solid", fgColor=sky)
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border
    # Right-side label column (mirror, like the AUGT1 sample)
    end_col = max_row + 2
    c = ws.cell(row=HDR_ROW, column=end_col, value="Rng")
    c.font = Font(bold=True, color=navy)
    c.fill = PatternFill("solid", fgColor=sky)
    c.alignment = Alignment(horizontal="center")
    c.border = cell_border
    ws.row_dimensions[HDR_ROW].height = 22

    # Data rows (ranges descending — top range at top)
    packet_by_pos = {(p["range_n"], p["row_n"]): p for p in packets}
    spike_boundaries = set()
    if layout:
        # Add a coloured stripe at row boundaries between spikes
        prev = None
        for spike_no, rows in sorted(layout.row_order_in_spike.items()):
            if not rows:
                continue
            if prev is not None:
                # boundary between this spike and previous
                spike_boundaries.add(min(rows))
            prev = spike_no

    for excel_row, rng in enumerate(range(max_range, 0, -1), start=HDR_ROW + 1):
        # Range label (left side)
        c = ws.cell(row=excel_row, column=1, value=rng)
        c.font = Font(bold=True, color=navy)
        c.fill = PatternFill("solid", fgColor=sky)
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border

        for fr in range(1, max_row + 1):
            excel_col = fr + 1
            p = packet_by_pos.get((rng, fr))
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=False)
            if p:
                value = p.get(value_key)
                if value:
                    cell.value = str(value)[:14]
                    cell.font = Font(size=9, color=navy)
            # Spike-boundary highlight on the LEFT edge of cells at boundary rows
            if fr in spike_boundaries:
                left = Side(style="medium", color="0678CD")
                cell.border = Border(top=thin, bottom=thin, left=left, right=thin)

        # Range label (right side mirror)
        c = ws.cell(row=excel_row, column=end_col, value=rng)
        c.font = Font(bold=True, color=navy)
        c.fill = PatternFill("solid", fgColor=sky)
        c.alignment = Alignment(horizontal="center")
        c.border = cell_border

    # Sizing
    ws.column_dimensions["A"].width = 9
    for col_idx in range(2, max_row + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    ws.column_dimensions[get_column_letter(end_col)].width = 7
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=2)


def _write_filtered_data_tab(ws, headers: list[str], packets: list[dict],
                             col_map: list[str | None],
                             filter_fn=None,
                             intro_note: str | None = None) -> None:
    """Generic filtered subset tab: header row + rows from packets[col_map]
    for packets where filter_fn(p) is truthy."""
    start_row = 1
    if intro_note:
        c = ws.cell(row=1, column=1, value=intro_note)
        c.font = Font(italic=True, color="A0522D")
        c.fill = PatternFill("solid", fgColor="FFF7E0")
        ws.merge_cells(start_row=1, end_row=1,
                       start_column=1, end_column=min(len(headers), 8))
        start_row = 3
    # Header
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=start_row, column=i, value=h)
        c.font = Font(bold=True, color="092A40", size=11)
        c.fill = PatternFill("solid", fgColor="E9F3FC")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="092A40"))
    ws.row_dimensions[start_row].height = 26

    # Data rows
    rows_added = 0
    for p in sorted(packets, key=lambda p: (p["range_n"], p["row_n"])):
        if filter_fn and not filter_fn(p):
            continue
        ws.append([(p.get(k) if k else None) for k in col_map])
        rows_added += 1

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    _autosize(ws, len(headers))
    return rows_added


def write_workbook(out_path: Path, nursery_code: str,
                   packets: list[dict],
                   nursery_name: str = "",
                   season: str = "",
                   breeder: str = "",
                   layout: MapLayout | None = None) -> None:
    """Produce the full multi-tab workbook from PRISM packets.

    Tabs match the original AUGT1-26S-IMI sample exactly (15 tabs):
      1. Nursery site         — raw PRISM data
      2. Nursery data         — nursery header info
      3. Field Map            — 2D grid (range × row) with Hybrid Code labels
      4. Material Map         — 2D grid (range × row) with Material ID labels
      5. Packet Prep ⭐       — 25-col table with QR CODE text + colored digits
      6. Nursery list         — Source IDs (with Hybrid Code) grouped by count
      7. Replacements done    — template (Qrcode, Replacement, Status)
      8. Planting error noted — template
      9. Fieldbook            — Range/Row/R_R/blanks/Material/Source/Gen/CMS/Comments
     10. BC0 labels           — BC* generations only, with TFMSA / Pollen columns
     11. Date recording       — recurrent (BC*, F*) packets with date columns
     12. Pulling bags         — same packets as Date recording, bag-pulling tracker
     13. Operations           — growth-stage template
     14. Comments             — free-text team notes template
     15. BC0 TFMSA record     — Day 7 / Day 10 / Day 13 TFMSA spray dates
     16. TFMSA Spray plots    — BC* plots that get TFMSA
     17. Hy Heights           — F1 hybrids only, height tracker
    """
    wb = Workbook()
    wb.remove(wb.active)

    # ── 1. Nursery site (raw PRISM, sorted by range/row) ──
    ns = wb.create_sheet("Nursery site")
    ns.append(PRISM_COLS)
    _format_header_row(ns)
    for p in sorted(packets, key=lambda p: (p["range_n"], p["row_n"])):
        ns.append([
            p["range_n"], p["row_n"],
            p.get("material_id"), p.get("inbred_code"), p.get("source_id"),
            p.get("cms"), p.get("generation"), p.get("comments"),
            p.get("pedigree"), p.get("hybrid_code"),
            p.get("trait_name"), p.get("plant_no"), p.get("loc_seq"),
            p.get("subseq_flag"), p.get("entry_book_project"),
            p.get("entry_book_name"), p.get("entry_no"),
        ])
    ns.freeze_panes = "A2"
    _autosize(ns, len(PRISM_COLS))

    # ── 2. Nursery data (header) ──
    nd = wb.create_sheet("Nursery data")
    nd["A1"] = "R&D Fieldbook - Grain Sorghum"
    nd["A1"].font = Font(bold=True, size=14, color="092A40")
    nd["A3"] = f"Nursery: {nursery_name or nursery_code}"
    nd["A4"] = f"Code:    {nursery_code}"
    if season:  nd["A5"] = f"Season:  {season}"
    if breeder: nd["A6"] = f"Breeder: {breeder}"
    nd.column_dimensions["A"].width = 60

    # ── 3. Map (2D grid: Hybrid Code per range/row — "Field Map" view) ──
    fm = wb.create_sheet("Map")
    _write_grid_map(fm, packets, value_key="hybrid_code",
                    title=f"Map — {nursery_code}",
                    layout=layout)

    # ── 4. Material Map (2D grid: Material ID per range/row) ──
    mm = wb.create_sheet("Material Map")
    _write_grid_map(mm, packets, value_key="material_id",
                    title=f"Material Map — {nursery_code}",
                    layout=layout)

    # ── 5. Packet Prep ⭐ (25 cols, exact sample-match) ──
    pp = wb.create_sheet("Packet Prep")
    pp.append(PACKET_PREP_COLS)
    _format_header_row(pp)
    sorted_packets = sorted(packets, key=lambda p: (p["spike"], p["rack_order"]))
    for p in sorted_packets:
        th, hu, te, on = rack_digits(p["rack_order"])
        pp.append([
            qr_text(p),
            p["range_n"], p["row_n"], p["plot"],
            p["spike"], p["rack_order"],
            th, hu, te, on,
            p.get("material_id"), p.get("inbred_code"),
            p.get("source_id"), p.get("cms"),
            p.get("generation"), p.get("comments"),
            p.get("pedigree"), p.get("hybrid_code"),
            p.get("trait_name"), p.get("plant_no"), p.get("loc_seq"),
            p.get("subseq_flag"), p.get("entry_book_project"),
            p.get("entry_book_name"), p.get("entry_no"),
        ])
    # Color-code the four digit columns (Thousands/Hundreds/Tens/Ones).
    for col_idx, name in enumerate(PACKET_PREP_COLS, start=1):
        if name in DIGIT_COLORS:
            color = DIGIT_COLORS[name]
            for row in pp.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for c in row:
                    c.font = Font(bold=True, color=color, size=11)
                    c.alignment = Alignment(horizontal="center")
    pp.freeze_panes = "B2"
    _autosize(pp, len(PACKET_PREP_COLS), default=12)
    pp.column_dimensions["A"].width = 38  # QR CODE column wider
    pp.column_dimensions["D"].width = 9   # Plot
    pp.column_dimensions["E"].width = 7   # SPIKE#
    pp.column_dimensions["F"].width = 11  # RACK ORDER
    for letter in ("G", "H", "I", "J"):
        pp.column_dimensions[letter].width = 6  # digit columns

    # ── 5. Nursery list (Source ID + Hybrid Code, with counts) ──
    nl = wb.create_sheet("Nursery list")
    nl.append(["", "", "Source ID (Hybrid Code)", "Repeats", "Qty Required",
               "Inbred Code", "Hybrid Code", "Notes"])
    _format_header_row(nl)
    qty_per_packet = 1.4
    groups: dict[str, dict] = {}
    for p in packets:
        sid = p.get("source_id") or "(unknown)"
        hc = p.get("hybrid_code")
        label = f"{sid} ({hc})" if hc else sid
        g = groups.setdefault(label, {
            "reps": 0,
            "inbred": p.get("inbred_code") or "",
            "hybrid": hc or "",
        })
        g["reps"] += 1
    for label in sorted(groups):
        g = groups[label]
        nl.append([None, None, label, g["reps"],
                   round(g["reps"] * qty_per_packet, 1),
                   g["inbred"], g["hybrid"], None])
    nl.freeze_panes = "A2"
    nl.column_dimensions["C"].width = 42
    for letter in ("D", "E"): nl.column_dimensions[letter].width = 12
    for letter in ("F", "G", "H"): nl.column_dimensions[letter].width = 18

    # ── 6. Fieldbook (10 cols, sorted serpentine by range then row) ──
    fb = wb.create_sheet("Fieldbook")
    fb.append(FIELDBOOK_COLS)
    _format_header_row(fb)
    for p in sorted(packets, key=lambda p: (p["range_n"], p["row_n"])):
        fb.append([
            p["range_n"], p["row_n"], p["plot"],
            None, None,                                # Crossed bags, Bagging Info
            p.get("material_id"), p.get("source_id"),
            p.get("generation"), p.get("cms"),
            p.get("comments"),
        ])
    fb.freeze_panes = "A2"
    _autosize(fb, len(FIELDBOOK_COLS))
    fb.page_setup.orientation = "landscape"
    fb.page_setup.fitToWidth = 1
    fb.print_options.horizontalCentered = True
    # print_title_rows expects a fully-qualified row range
    fb.print_title_rows = "$1:$1"

    # ── 7. Replacements done (template, ready for PRISM upload) ──
    rd = wb.create_sheet("Replacements done")
    rd["A1"] = ("Rename this tab to 'Replacements done' once Breeder has "
                "updated PRISM.")
    rd["A1"].font = Font(italic=True, color="A0522D")
    rd["A1"].fill = PatternFill("solid", fgColor="FFF7E0")
    rd.append([])
    rd.append(["Qrcode", "Replacement", "Status", "Notes"])
    _format_header_row(rd, row_idx=3)
    rd.column_dimensions["A"].width = 40
    rd.column_dimensions["B"].width = 32
    rd.column_dimensions["C"].width = 12
    rd.column_dimensions["D"].width = 28

    # ── 8. Planting error noted (template) ──
    pe = wb.create_sheet("Planting error noted")
    pe["A1"] = ("Rename this tab to 'Planting error noted' once Breeder has "
                "updated PRISM.")
    pe["A1"].font = Font(italic=True, color="A0522D")
    pe["A1"].fill = PatternFill("solid", fgColor="FFF7E0")
    pe.append([])
    pe.append(["Plot", "Range", "Row", "Description",
               "Severity", "Date noticed", "Status"])
    _format_header_row(pe, row_idx=3)
    for col, w in zip("ABCDEFG", (12, 8, 8, 40, 12, 14, 12)):
        pe.column_dimensions[col].width = w

    # ── 10. BC0 labels (BC* packets only — TFMSA / Pollen tracking) ──
    bc = wb.create_sheet("BC0 labels")
    bc_headers = ["Range", "Row", "Crossed bags", "TFMSA", "Pollen",
                  "TFMSA/Pollen", "Nursery Name", "Bagging Info", "Material ID",
                  "Source ID", "Gen", "CMS", "Comments"]
    bc_cols = ["range_n", "row_n", None, None, None, None, None, None,
               "material_id", "source_id", "generation", "cms", "comments"]
    n_bc = _write_filtered_data_tab(bc, bc_headers, packets, bc_cols,
                                    filter_fn=lambda p: _is_bc_generation(p.get("generation")))
    # Stamp the Nursery Name column with the nursery code
    if n_bc:
        nn_col = bc_headers.index("Nursery Name") + 1
        for r in range(2, n_bc + 2):
            bc.cell(row=r, column=nn_col, value=nursery_code)

    # ── 11. Date recording (recurrent BC*/F* packets) ──
    dr = wb.create_sheet("Date recording")
    dr_headers = ["Range", "Row", "Plot", "1", "2",
                  "Material ID", "Source ID", "Gen", "CMS", "Comments"]
    dr_cols = ["range_n", "row_n", "plot", None, None,
               "material_id", "source_id", "generation", "cms", "comments"]
    _write_filtered_data_tab(dr, dr_headers, packets, dr_cols,
                             filter_fn=lambda p: _is_recurrent(p.get("generation")))

    # ── 12. Pulling bags (same population as Date recording) ──
    pb = wb.create_sheet("Pulling bags")
    _write_filtered_data_tab(pb, dr_headers, packets, dr_cols,
                             filter_fn=lambda p: _is_recurrent(p.get("generation")))

    # ── 13. Operations (growth-stage template) ──
    ops = wb.create_sheet("Operations")
    ops["A1"] = "Operations Overview"
    ops["A1"].font = Font(bold=True, size=14, color="092A40")
    ops.merge_cells("A1:F1")
    ops.append([])
    ops.append(["Stage", "Plan", "Reminders", "Comments"])
    _format_header_row(ops, row_idx=3)
    for i, (stage, rgb) in enumerate(GROWTH_STAGES, start=4):
        c = ops.cell(row=i, column=1, value=stage)
        c.font = Font(bold=True, color="092A40")
        c.fill = PatternFill("solid", fgColor="{:02X}{:02X}{:02X}".format(*rgb))
        c.alignment = Alignment(horizontal="center", vertical="center")
        ops.row_dimensions[i].height = 30
    ops.column_dimensions["A"].width = 16
    for col in "BCD":
        ops.column_dimensions[col].width = 36

    # ── 14. Comments (free-text team notes) ──
    cm = wb.create_sheet("Comments")
    cm["A1"] = "Team Comments"
    cm["A1"].font = Font(bold=True, size=14, color="092A40")
    cm.merge_cells("A1:D1")
    cm.append([])
    cm.append(["Date", "Tech", "Topic", "Comment"])
    _format_header_row(cm, row_idx=3)
    for col, w in zip("ABCD", (14, 12, 18, 60)):
        cm.column_dimensions[col].width = w

    # ── 15. BC0 TFMSA record (Day 7 / Day 10 / Day 13 spray-dates template) ──
    tr = wb.create_sheet("BC0 TFMSA record")
    tr["A1"] = "BC0 TFMSA record"
    tr["A1"].font = Font(bold=True, size=14, color="092A40")
    tr["A2"] = ("Record TFMSA spray dates and crossing observations here.")
    tr["A2"].font = Font(italic=True, color="5A7896")
    tr.merge_cells("A1:N1"); tr.merge_cells("A2:N2")
    # Header row at row 3: Day 7, Day 10, Day 13 (each spans 5 cols like sample)
    headers_with_span = [("Day 7", 1), ("Day 10", 6), ("Day 13", 11)]
    for label, start_col in headers_with_span:
        c = tr.cell(row=3, column=start_col, value=label)
        c.font = Font(bold=True, color="092A40", size=12)
        c.fill = PatternFill("solid", fgColor="E9F3FC")
        c.alignment = Alignment(horizontal="center")
        tr.merge_cells(start_row=3, end_row=3,
                       start_column=start_col, end_column=start_col + 4)
    # Sub-header at row 4: Range, Row, Source ID, CMS, Notes (repeated 3x)
    for block_idx in range(3):
        start_col = 1 + block_idx * 5
        for offset, h in enumerate(["Range", "Row", "Source ID", "CMS", "Notes"]):
            c = tr.cell(row=4, column=start_col + offset, value=h)
            c.font = Font(bold=True, color="092A40", size=10)
            c.fill = PatternFill("solid", fgColor="F3F7FB")
            c.alignment = Alignment(horizontal="center")
    tr.freeze_panes = "A5"
    for i in range(1, 16):
        tr.column_dimensions[get_column_letter(i)].width = 11

    # ── 16. TFMSA Spray plots (BC* plots that get TFMSA) ──
    sp = wb.create_sheet("TFMSA Spray plots")
    sp["A1"] = "TFMSA Spray plots"
    sp["A1"].font = Font(bold=True, size=14, color="092A40")
    sp.merge_cells("A1:E1")
    sp.append([])
    sp_headers = ["Range", "Row", "Source ID", "CMS", "Gen"]
    sp_cols = ["range_n", "row_n", "source_id", "cms", "generation"]
    # Pre-write header at row 3
    for i, h in enumerate(sp_headers, start=1):
        c = sp.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="092A40")
        c.fill = PatternFill("solid", fgColor="E9F3FC")
        c.alignment = Alignment(horizontal="center")
        c.border = Border(bottom=Side(style="medium", color="092A40"))
    sp.row_dimensions[3].height = 24
    # Filtered data rows
    for p in sorted(packets, key=lambda p: (p["range_n"], p["row_n"])):
        if _is_bc_generation(p.get("generation")):
            sp.append([p.get(k) for k in sp_cols])
    sp.freeze_panes = "A4"
    _autosize(sp, len(sp_headers))

    # ── 17. Hy Heights (F1 hybrids — height tracking) ──
    hh = wb.create_sheet("Hy Heights")
    hh_headers = ["Range", "Row", "Height in CM", "Material ID",
                  "Source ID", "Gen", "CMS"]
    hh_cols = ["range_n", "row_n", None, "material_id",
               "source_id", "generation", "cms"]
    _write_filtered_data_tab(hh, hh_headers, packets, hh_cols,
                             filter_fn=lambda p: _is_hybrid_f1(p.get("generation")))

    out_path.parent.mkdir(exist_ok=True, parents=True)
    wb.save(out_path)


def init_nursery(input_path: Path, nursery_code: str,
                 sheet: str = "Sheet1",
                 nursery_name: str = "",
                 season: str = "",
                 breeder: str = "") -> dict:
    """Run the full nursery initialisation pipeline. Callable from API code."""
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    packets = read_prism_export(input_path, sheet)
    layout = parse_map(input_path)
    assign_spike_rack(packets, layout)
    conn = init_db()
    insert_packets(conn, nursery_code, packets)
    wb_path = OUT_DIR / f"{nursery_code}_workbook.xlsx"
    write_workbook(wb_path, nursery_code, packets,
                   nursery_name=nursery_name, season=season, breeder=breeder,
                   layout=layout)
    return {
        "nursery_code": nursery_code,
        "packet_count": len(packets),
        "spike_count": (layout.spike_count if layout else 0),
        "map_source": (layout.sheet_name if layout else "(no Map found — serpentine fallback)"),
        "workbook_path": str(wb_path),
    }


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, type=Path,
                    help="PRISM export .xlsx")
    ap.add_argument("--nursery-code", required=True,
                    help="Short nursery code, e.g. AUGT1-26S-IMI")
    ap.add_argument("--sheet", default="Sheet1", help="Sheet name in the export")
    ap.add_argument("--name", default="", help="Full nursery name")
    ap.add_argument("--season", default="", help="Season label, e.g. 2026S")
    ap.add_argument("--breeder", default="", help="Lead breeder name")
    args = ap.parse_args()

    print(f"Reading PRISM export {args.input} (sheet={args.sheet})…")
    result = init_nursery(args.input, args.nursery_code, args.sheet,
                          nursery_name=args.name, season=args.season,
                          breeder=args.breeder)
    print(f"  → {result['packet_count']} packets")
    print(f"  → Map: {result['map_source']} ({result['spike_count']} spike(s))")
    print(f"  → wrote {result['workbook_path']}")
    print("Done.")


if __name__ == "__main__":
    main()
